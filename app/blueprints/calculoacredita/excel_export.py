import io
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .services import CATEGORIA_GER, CATEGORIA_PRE, CATEGORIA_APO

FACTOR_E, FACTOR_P, FACTOR_R = 1, 2, 3


def _ruta_modelo_excel():
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    nombre_archivo = "modelo excel acredita.xlsx"
    for carpeta in ("recursos", "archivos"):
        ruta = os.path.join(base, carpeta, nombre_archivo)
        if os.path.exists(ruta):
            return ruta
    ruta_blueprint = os.path.join(os.path.dirname(os.path.abspath(__file__)), "plantillas", nombre_archivo)
    if os.path.exists(ruta_blueprint):
        return ruta_blueprint
    return os.path.join(base, "recursos", nombre_archivo)


def generar_excel_CalculoAcredita(datos):
    ruta_modelo = _ruta_modelo_excel()
    if not os.path.exists(ruta_modelo):
        return _generar_excel_sin_modelo(datos)

    wb = load_workbook(ruta_modelo, data_only=False)
    ws = wb.active

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= 33 and merged_range.max_row >= 8:
            try:
                ws.unmerge_cells(str(merged_range))
            except Exception:
                pass

    result = datos["result"]
    datos_categoria = {d["categoria"]: d for d in datos["datos_categoria"]}
    porcentaje_final = datos["porcentaje_final"]
    ipress = datos["ipress"]
    autoevaluacion = datos["autoevaluacion"]
    nombres_macroproceso = datos["nombres_macroproceso"]

    ws["D1"] = f"{ipress.codigo_ipress}-{ipress.nombre_ipress}"
    ws["D2"] = ipress.nivel_ipress or "I-3"
    ws["D3"] = "AutoEvaluación"
    ws["D4"] = getattr(autoevaluacion, "periodo", 2025)
    ws["D5"] = 1

    cat_macros = {"GER": CATEGORIA_GER, "PRE": CATEGORIA_PRE, "APO": CATEGORIA_APO}
    row = 8

    for cod_cate in ["GER", "PRE", "APO"]:
        macros = cat_macros[cod_cate]
        cat_info = datos_categoria.get(cod_cate, {})

        for idx, codmacro in enumerate(macros):
            if codmacro not in result:
                continue
            r = result[codmacro]
            nombre_mp = nombres_macroproceso.get(codmacro, codmacro)

            ws.cell(row, 1, cod_cate if idx == 0 else "") #A
            ws.cell(row, 2, codmacro) #B
            ws.cell(row, 3, nombre_mp) #C

            e_val, p_val, g_val = r.get("E_valor_estructura", 0), r.get("F_valor_proceso", 0), r.get("G_valor_resultado", 0)
            h_val = r.get("H_cantidad_criterios_reportados", 0)
            ws.cell(row, 4, h_val) #D
            ws.cell(row, 5, e_val) #E
            ws.cell(row, 6, p_val) #F
            ws.cell(row, 7, g_val) #G
            ws.cell(row, 8, h_val) #H
            #ws.cell(row, 7, max(0, h_val - e_val - p_val - g_val) if h_val else 0)

            i_val = r.get("I_puntaje_maximo_estructura", 0)
            j_val = r.get("J_puntaje_maximo_proceso", 0)
            k_val = r.get("K_puntaje_maximo_resultado", 0)
            l_val = r.get("L_sumatoria_total_pmt", 0)
            #t_crit = max(0, h_val - e_val - p_val - g_val) if h_val else 0
            ws.cell(row, 9, i_val) #I
            ws.cell(row, 10, j_val) #J
            ws.cell(row, 11, k_val) #K
           # ws.cell(row, 12, t_crit * 2 if t_crit else 0)
            ws.cell(row, 12, l_val) #L
            

            ws.cell(row, 13, FACTOR_E) #M
            ws.cell(row, 14, FACTOR_P) #N
            ws.cell(row, 15, FACTOR_R) #O
            ws.cell(row, 16, FACTOR_E + FACTOR_P + FACTOR_R) #P

            ws.cell(row, 17, r.get("Q_puntaje_maximo_estructura_peso", 0))  #Q
            ws.cell(row, 18, r.get("R_puntaje_maximo_proceso_peso", 0)) #R
            ws.cell(row, 19, r.get("S_puntaje_maximo_resultado_peso", 0))  #S
            ws.cell(row, 20, r.get("T_sumatoria_total", 0)) #T

            u, v, w = r.get("U_ptc_estructura_criterio", 0), r.get("V_ptc_proceso_criterio", 0), r.get("W_ptc_resultado_criterio", 0)
            ws.cell(row, 21, round(u, 2)) #U
            ws.cell(row, 22, round(v, 2)) #V
            ws.cell(row, 23, round(w, 2)) #W
            #ws.cell(row, 24, round(u + v + w, 2) if (u + v + w) else 0)
            ws.cell(row, 24, l_val) #X
            ws.cell(row, 25, u + v+ w) #Y
            
            z = r.get("Z_puntaje_obtenido_por_tipo_macroproceso_estructura", 0)
            aa = r.get("AA_puntaje_obtenido_por_tipo_macroproceso_proceso", 0)
            ab = r.get("AB_puntaje_obtenido_por_tipo_macroproceso_resultado", 0)
            ws.cell(row, 26, z)
            ws.cell(row, 27, aa)
            ws.cell(row, 28, ab)
            ws.cell(row, 29, z + aa + ab)

            ad = r.get("AD_ptje_max_proporcion_estructura", 0)
            ae = r.get("AE_ptje_max_proporcion_proceso", 0)
            af = r.get("AF_ptje_max_proporcion_resultado", 0)
            ws.cell(row, 30, round(ad, 2))
            ws.cell(row, 31, round(ae, 2))
            ws.cell(row, 32, round(af, 2))
            ws.cell(row, 33, round(ad + ae + af, 2))

            ws.cell(row, 34, round(r.get("AH_ptje_obtenido_mp_estructura", 0), 2))
            ws.cell(row, 35, round(r.get("AI_ptje_obtenido_mp_proceso", 0), 2))
            ws.cell(row, 36, round(r.get("AJ_ptje_obtenido_mp_resultado", 0), 2))
            ws.cell(row, 37, round(r.get("AK_sumatoria_puntaje_obtenido_mp_todos", 0), 2))

            ws.cell(row, 38, r.get("AL_porcentaje_macroproceso_constante", 0))
            ws.cell(row, 39, round(r.get("AM_puntaje_min_por_macroproceso", 0), 2))
            ws.cell(row, 40, round(r.get("AN_puntaje_obtenido_por_macroproceso", 0), 2))
            ws.cell(row, 41, round(r.get("AO_cumplimiento_macroproceso", 0), 2))

            if idx == 0:
                ap = cat_info.get("AP_puntaje_max_categoria", 0)
                aq = cat_info.get("AQ_porcentaje_categoria", 0)
                ar = cat_info.get("AR_total_ptje_categoria_max", 0)
                as_cat = cat_info.get("AS_puntaje_max_categoria", 0)
                at = cat_info.get("AT_puntaje_obtenido", 0)
                cumpl_cat = round((at / ar * 100), 2) if ar and ar > 0 else 0
                ws.cell(row, 42, round(ap, 2))
                ws.cell(row, 43, round(aq, 2))
                ws.cell(row, 44, round(ar, 2))
                ws.cell(row, 45, round(as_cat, 2))
                ws.cell(row, 46, round(at, 2))
                ws.cell(row, 47, round(cumpl_cat, 2))

            row += 1

    tot_e_cr = sum(result.get(m, {}).get("E_valor_estructura", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_p_cr = sum(result.get(m, {}).get("F_valor_proceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_r_cr = sum(result.get(m, {}).get("G_valor_resultado", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_h = sum(result.get(m, {}).get("H_cantidad_criterios_reportados", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_i = sum(result.get(m, {}).get("I_puntaje_maximo_estructura", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_j = sum(result.get(m, {}).get("J_puntaje_maximo_proceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_k = sum(result.get(m, {}).get("K_puntaje_maximo_resultado", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_l = sum(result.get(m, {}).get("L_sumatoria_total_pmt", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_q = sum(result.get(m, {}).get("Q_puntaje_maximo_estructura_peso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_r = sum(result.get(m, {}).get("R_puntaje_maximo_proceso_peso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_s = sum(result.get(m, {}).get("S_puntaje_maximo_resultado_peso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_t = sum(result.get(m, {}).get("T_sumatoria_total", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_z = sum(result.get(m, {}).get("Z_puntaje_obtenido_por_tipo_macroproceso_estructura", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_aa = sum(result.get(m, {}).get("AA_puntaje_obtenido_por_tipo_macroproceso_proceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_ab = sum(result.get(m, {}).get("AB_puntaje_obtenido_por_tipo_macroproceso_resultado", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_ah = sum(result.get(m, {}).get("AH_ptje_obtenido_mp_estructura", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_ai = sum(result.get(m, {}).get("AI_ptje_obtenido_mp_proceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_aj = sum(result.get(m, {}).get("AJ_ptje_obtenido_mp_resultado", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_ak = sum(result.get(m, {}).get("AK_sumatoria_puntaje_obtenido_mp_todos", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_am = sum(result.get(m, {}).get("AM_puntaje_min_por_macroproceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    tot_an = sum(result.get(m, {}).get("AN_puntaje_obtenido_por_macroproceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    ##tot_ao = round((tot_an / tot_am * 100), 2) if tot_am else 0
    ##t_crit_total = max(0, tot_h - tot_e_cr - tot_p_cr - tot_r_cr)
   
    ws.cell(row, 1, "")
    ws.cell(row, 2, "")
    ws.cell(row, 3, "Total")
    ws.cell(row, 4, tot_e_cr +tot_p_cr + tot_r_cr)
    ws.cell(row, 5, tot_e_cr) #E
    ws.cell(row, 6, tot_p_cr) #F
    ws.cell(row, 7, tot_r_cr) #G
    ws.cell(row, 8, tot_e_cr +tot_p_cr + tot_r_cr)  #H
    ws.cell(row, 9, tot_i)
    ws.cell(row, 10, tot_j)
    ws.cell(row, 11, tot_k)
    ws.cell(row, 12, tot_i + tot_j +tot_k) #L
    ws.cell(row, 13, FACTOR_E*22)
    ws.cell(row, 14, FACTOR_P*22)
    ws.cell(row, 15, FACTOR_R*22)
    ws.cell(row, 16, (FACTOR_E + FACTOR_P + FACTOR_R)*22)
    ws.cell(row, 17, tot_q)
    ws.cell(row, 18, tot_r)
    ws.cell(row, 19, tot_s)
    ws.cell(row, 20, tot_t)
    u_sum = sum(result.get(m, {}).get("U_ptc_estructura_criterio", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    v_sum = sum(result.get(m, {}).get("V_ptc_proceso_criterio", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    w_sum = sum(result.get(m, {}).get("W_ptc_resultado_criterio", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    ws.cell(row, 21, round(u_sum, 2))
    ws.cell(row, 22, round(v_sum, 2))
    ws.cell(row, 23, round(w_sum, 2))
    ws.cell(row, 24, round(tot_i + tot_j +tot_k)) #X
    ws.cell(row, 25, u_sum + v_sum + w_sum) #Y
    ws.cell(row, 26, tot_z)
    ws.cell(row, 27, tot_aa)
    ws.cell(row, 28, tot_ab)
    ws.cell(row, 29, tot_z + tot_aa + tot_ab) #AC
    
    ad_sum = sum(result.get(m, {}).get("AD_ptje_max_proporcion_estructura", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    ae_sum = sum(result.get(m, {}).get("AE_ptje_max_proporcion_proceso", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    af_sum = sum(result.get(m, {}).get("AF_ptje_max_proporcion_resultado", 0) for m in cat_macros["GER"] + cat_macros["PRE"] + cat_macros["APO"] if m in result)
    ws.cell(row, 30, round(ad_sum, 2)) #AD
    ws.cell(row, 31, round(ae_sum, 2))
    ws.cell(row, 32, round(af_sum, 2))
    ws.cell(row, 33, round(ad_sum + ae_sum + af_sum, 2))
    ws.cell(row, 34, round(tot_ah, 2))
    ws.cell(row, 35, round(tot_ai, 2))
    ws.cell(row, 36, round(tot_aj, 2))
    ws.cell(row, 37, round(tot_ak, 2))
    ws.cell(row, 38, 100) #AL
    ws.cell(row, 39, round(tot_am, 2))
    ws.cell(row, 40, round(tot_an, 2))
    ws.cell(row, 41, "") #AO
    ar_total_ar = sum(d.get("AR_total_ptje_categoria_max", 0) for d in datos_categoria.values())
    at_total = sum(d.get("AT_puntaje_obtenido", 0) for d in datos_categoria.values())
    ws.cell(row, 42, round(at_total, 2)) #AP
    ws.cell(row, 43, 100) #AQ
    ws.cell(row, 44, round(ar_total_ar, 2)) #AR
    ws.cell(row, 45, round(tot_an, 2)) #AS
    ws.cell(row, 46, round(at_total, 2)) #AT
    ws.cell(row, 47, "") #AU
    

    bloques_categoria = [(8, 13), (14, 19), (20, 29)]
    for start_row, end_row in bloques_categoria:
        for col in range(42, 48):
            try:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col
                )
            except Exception:
                pass

    puntaje_row = 33
    ws.cell(puntaje_row, 45, "Puntaje Final (%)")
    ws.cell(puntaje_row, 46, round(porcentaje_final, 2))

    def _clear_cell(r, c):
        try:
            cell = ws.cell(row=r, column=c)
            cell.value = None
        except Exception:
            pass
    _clear_cell(5, 34)
    _clear_cell(33, 33)
    _clear_cell(33, 39)
    sin_fill = PatternFill(fill_type=None)
    for col in range(1, 45):
        try:
            ws.cell(33, col).fill = sin_fill
        except Exception:
            pass
    for col in range(47, 50):
        try:
            ws.cell(33, col).fill = sin_fill
        except Exception:
            pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _generar_excel_sin_modelo(datos):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cálculo Acreditación"

    ipress = datos["ipress"]
    ws["A1"] = "Establecimiento"
    ws["B1"] = f"{ipress.codigo_ipress}-{ipress.nombre_ipress}"
    ws["A2"] = "Tipo"
    ws["B2"] = ipress.nivel_ipress or "I-3"
    ws["A3"] = "Puntaje Final (%)"
    ws["B3"] = round(datos["porcentaje_final"], 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
